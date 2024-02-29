from flask import Flask,request,jsonify,json
from flask_sqlalchemy import SQLAlchemy
import uuid # for public id
from werkzeug.security import generate_password_hash, check_password_hash
# imports for PyJWT authentication
import jwt
from datetime import datetime, timedelta
from functools import wraps

from flask_sqlalchemy import SQLAlchemy

from flask_jwt_extended import create_access_token ,JWTManager,get_jwt_identity,jwt_required
from flask_restful import Api,Resource
from openpyxl import load_workbook

app=Flask(__name__)
app.config['SECRET_KEY']='thissecretkey'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sey.db'
api=Api(app)
db=SQLAlchemy(app)
jwt=JWTManager(app)


class User(db.Model):
    id=db.Column(db.Integer(),primary_key=True)
    username=db.Column(db.String(200))
    password=db.Column(db.String(200))
    
@app.route('/signup', methods =['POST'])
def signup():
        data=request.get_json()
        username=data['username']
        password=data['password']

        if not username and not password:
            return jsonify({'message':'username and password are invalid'}),400
        if User.query.filter_by(username=username).first():
            return jsonify({'message':'username allredy exsit'})
        new_user=User(username=username,password=password)
        db.session.add(new_user)
        db.session.commit()
        return jsonify({'message':'new user created sucsessfully'})
@app.route('/login', methods =['POST'])
def login():
        data = request.get_json()
        username = data['username']
        password = data['password']

        user = User.query.filter_by(username=username).first()
        print(user)
        if user and user.password==password :
            access_token = create_access_token(identity=user.id)
            print(access_token)
            return {'access_token': access_token}, 200

        return {'message': 'Invalid credentials'}, 401



# # file parcel table 
class Std(db.Model):
    id=db.Column(db.Integer, primary_key=True,autoincrement=True)
    name=db.Column(db.String(100), nullable=False)
    age=db.Column(db.Integer(), nullable=False)   


# decorator for verifying the JWT
def token_required(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		token = None
		if 'x-access-token' in request.headers:
			token = request.headers['x-access-token']
		if not token:
			return jsonify({'message' : 'Token is missing !!'}), 401

		try:
			data = jwt.decode(token, app.config['SECRET_KEY'])
			current_user = User.query\
				.filter_by(public_id = data['public_id'])\
				.first()
		except:
			return jsonify({
				'message' : 'Token is invalid !!'
			}), 401
		return f(current_user, *args, **kwargs)

	return decorated

@app.route('/user', methods =['GET'])
@token_required
def get_all_users(current_user):
	users = User.query.all()
	return ('helloo')



@token_required
@app.route('/fileparser',methods=['POST'])
def fileparser():
    data=request.files['file']
    workbook=load_workbook(data)

    print(workbook)
    sheet=workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
        name,age=row
        data2=Std(name=name,age=age)
        db.session.add(data2)

        db.session.commit()
    

    return 'msg: file uploaded Successfully '

# '''Crud on Flask '''

@app.route('/create', methods=['POST'])
def create():
    if request.method=='POST':
        name=request.json['name']
        age=request.json['age']
        new=Std(name=name,age=age)
        db.session.add(new)
        db.session.commit()


    return 'user created successfully '

@app.route('/read', methods=['GET'])
def read():
    all_data = Std.query.all()
    
    # Convert SQLAlchemy objects to a list of dictionaries
    data_list = []
    for data in all_data:
        data_list.append({
            'id': data.id,
            "name":data.name,
            "age":data.age
            # Add other fields as needed
        })
    
    return jsonify(data_list)   


@app.route('/update', methods=['PUT'])
def update():
    if request.method=='PUT':
        id=request.json['id']
        name=request.json['name']
        age=request.json['age']
        new_data=Std.query.filter_by(id=id).first()
        new_data.name=name
        new_data.age=age
        db.session.add(new_data)
        db.session.commit()
        
    return "data updated successfully.....!"    



@app.route('/delete', methods=['DELETE'])
def delete():
    if request.method=='DELETE':
        id=request.json['id']
        new=Std.query.filter_by(id=id).first()
        db.session.delete(new)
        db.session.commit()
    
    return "data deleted "


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # db.drop_all()
    app.run(debug=True)









